#把所有表格为文件夹放在一个目录下
import os
from openpyxl import load_workbook
wball = load_workbook('C:\\Users\\thril\\Desktop\\EXCELwork\\allwb.xlsx')
sheetall = wball['Sheet1']
columnmax = sheetall.max_column


def Findsheet(path,findstr):
    filenames = os.listdir(path)
    for i, filename in enumerate(filenames):
        # print('start find in sheets')
        wb = load_workbook(path+'\\'+filename)
        # 获取workbook中所有的表格
        sheets = wb.sheetnames
        # 循环遍历所有sheet
        for i in range(len(sheets)):
            findresult=sheets[i].find(findstr)
            if findresult != -1:
                # print('finded one sheet')
                tagetsheet = wb[sheets[i]]
                return tagetsheet
            else:
                continue

path='C:\\Users\\thril\\Desktop\\EXCELwork\\data'
#rowindex from 2 to 293
for rowindex in range(2,293):  
    namecell = sheetall['A'+str(rowindex)]
    namelong=len(namecell.value)-1
    shortname=namecell.value[0:namelong]
    print('start row_'+str(rowindex))
    tagetsheet=Findsheet(path,shortname)
    if tagetsheet is not None:
        for r in range(1, tagetsheet.max_row+1):
        #for r in range(1, 10): 
            for c in range(1, tagetsheet.max_column+1):
            #for c in range(1, 7):
                v = tagetsheet.cell(row=r, column=c).value
                if v == tagetsheet.title:
                    # print('finded_')
                    # print(tagetsheet.title)
                    fieldstr=tagetsheet.cell(row=r, column=2)
                    for index in range(c+1,tagetsheet.max_column+1):
                        fieldvalue=str(tagetsheet.cell(row=r, column=2).value)[0:4]+'_'+str(tagetsheet.cell(row=2, column=index).value)+'('+str(tagetsheet.cell(row=4, column=index).value)+')'
                        # print('fieldstr is '+fieldvalue)
                        #加入到总表中
                        columnmax=columnmax+1
                        sheetall.insert_cols(columnmax)
                        #写入表头
                        fieldcell=sheetall.cell(row=1, column=columnmax)
                        fieldcell.value = fieldvalue
                        #写入值
                        tagetcell=sheetall.cell(row=rowindex, column=columnmax)
                        tagetcell.value = tagetsheet.cell(row=r, column=index).value
                        # print('value is '+str(tagetcell.value))
    print('over row_'+str(rowindex))
wball.save('C:\\Users\\thril\\Desktop\\EXCELwork\\allwbafer.xlsx')
print('ok')
                






