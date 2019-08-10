# -*- coding: utf-8 -*-
#张老师安排，将黄河流域各县的数据整理到一个表格中
import sys
stdi,stdo,stde=sys.stdin,sys.stdout,sys.stderr 
reload(sys)
sys.stdin,sys.stdout,sys.stderr=stdi,stdo,stde 
sys.setdefaultencoding('utf-8')
import os
from openpyxl import load_workbook
wball = load_workbook('/home/JiQiulei/EXCELwork/allwb.xlsx')
sheetall = wball['Sheet1']
columnmax = sheetall.max_column


def Findsheet(path,findstr):
    filenames = os.listdir(path)
    for i, filename in enumerate(filenames):
        # print('start find in sheets')
        wb = load_workbook(path+'/'+filename)
        sheets = wb.sheetnames
        for i in range(len(sheets)):
            findresult=sheets[i].find(findstr)
            if findresult != -1:
                # print('finded one sheet')
                tagetsheet = wb[sheets[i]]
                return tagetsheet
            else:
                continue

path='/home/JiQiulei/EXCELwork/data'
#rowindex from 2 to 293
for rowindex in range(2,293):  
    namecell = sheetall['A'+str(rowindex)]
    namelong=len(namecell.value)-1
    shortname=namecell.value[0:namelong]
    #print('start row_'+str(rowindex))
    tagetsheet=Findsheet(path,shortname)
    if tagetsheet is not None:
        for r in range(1, tagetsheet.max_row+1):
        #for r in range(1, 10): 
            for c in range(1, tagetsheet.max_column+1):
            #for c in range(1, 7):
                v = tagetsheet.cell(row=r, column=c).value
                if v == tagetsheet.title:
                    # print('finded_')
                    # print(tagetsheet.title)
                    for index in range(c+1,tagetsheet.max_column+1):
                        fieldvalue=''
                        if c-2>=1:
                            fieldvalue=str(tagetsheet.cell(row=r, column=c-2).value)[0:4]+'_'+str(tagetsheet.cell(row=r, column=c-1).value)[0:4]+'_'+str(tagetsheet.cell(row=2, column=index).value)+'('+str(tagetsheet.cell(row=4, column=index).value)+')'
                        else:
                            fieldvalue=str(tagetsheet.cell(row=r, column=c-1).value)[0:4]+'_'+str(tagetsheet.cell(row=2, column=index).value)+'('+str(tagetsheet.cell(row=4, column=index).value)+')'                           
                        fieldvalue=fieldvalue.replace('_?','')                  
                        # print('fieldstr is '+fieldvalue)
                        findornot=False
                        for clonumber in range(1,columnmax):
                            alfield=sheetall.cell(row=1, column=clonumber)
                            if fieldvalue==alfield.value:
                                tagetcell=sheetall.cell(row=rowindex, column=clonumber)
                                tagetcell.value = tagetsheet.cell(row=r, column=index).value
                                findornot=True
                                break
                        if findornot==False:
                            columnmax=columnmax+1
                            sheetall.insert_cols(columnmax)
                            fieldcell=sheetall.cell(row=1, column=columnmax)
                            fieldcell.value = fieldvalue
                            tagetcell=sheetall.cell(row=rowindex, column=columnmax)
                            tagetcell.value = tagetsheet.cell(row=r, column=index).value
                            # print('value is '+str(tagetcell.value))
    print 'over row_'+str(rowindex)
wball.save('/home/JiQiulei/EXCELwork/allwbafer.xlsx')
print 'ok'
                






